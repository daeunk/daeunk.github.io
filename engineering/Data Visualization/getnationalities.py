# -*- coding: utf-8 -*-
from openpyxl import Workbook, load_workbook
import urllib
import json
# getting wikidata on actors' nationalities
new_wb = Workbook()
new_ws = new_wb.active

old_wb = load_workbook(filename = 'newactors.xlsx')
old_ws = old_wb['actors']

actor_url = 'http://www.wikidata.org/w/api.php?action=wbgetentities&sites=ptwiki&format=json&titles='
nationality_url = 'http://www.wikidata.org/w/api.php?action=wbgetentities&format=json&ids='

row_number = 1
for row in old_ws.iter_rows('A{}:A{}'.format(old_ws.min_row,old_ws.max_row)):
    for cell in row:
    	actor_name = cell.value
    	# print actor_name
    	new_ws.cell(row=row_number, column=1).value = actor_name
    	title = ''
    	for char in actor_name:
    		if char == ' ':
    			title += '_'
    		else:
    			title += char
    	title = title.encode('ascii', 'ignore').decode('ascii')
    	actor_final_url = actor_url + title
    	# actor_final_url = urllib.quote(actor_final_url.encode('utf-8'), ':/')
    	print str(row_number) + ": " + actor_final_url
    	# actor_page = urllib.urlopen(actor_final_url)
    	# actor_content = actor_page.read()
    	# actor_data = json.loads(actor_content)
    	actor_data = json.load(urllib.urlopen(actor_final_url))
    	entities = actor_data['entities']
    	# print actor_content
    	if 'claims' in entities[entities.keys()[0]].keys() and 'P27' in entities[entities.keys()[0]]['claims'].keys():
	    	nationality_id = entities[entities.keys()[0]]['claims']['P27'][0]['mainsnak']['datavalue']['value']['id']
	    	nationality_data = json.load(urllib.urlopen(nationality_url + nationality_id))
	    	ent = nationality_data['entities']
	    	nationality = ent[ent.keys()[0]]['labels']['en']['value']
	    	# print nationality
	    	new_ws.cell(row=row_number, column=2).value = nationality
    	row_number += 1
    	new_wb.save('ActorsNationality.xlsx')