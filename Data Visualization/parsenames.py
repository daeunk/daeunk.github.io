# -*- coding: utf-8 -*-
from openpyxl import Workbook, load_workbook
import urllib
import json
# Parsing actor list to create a new excel sheet of actor names
new_wb = Workbook()
new_ws = new_wb.active

old_wb = load_workbook(filename = 'actors.xlsx')
old_ws = old_wb['actors']

row_number = 1
for row in old_ws.iter_rows('A{}:A{}'.format(old_ws.min_row,old_ws.max_row)):
    for cell in row:
    	arr = cell.value.split(', ')
    	for entry in arr:
    		if (entry != ''):
	    		new_ws.cell(row=row_number, column=1).value = entry
	    		row_number += 1

new_wb.save('newactors.xlsx')