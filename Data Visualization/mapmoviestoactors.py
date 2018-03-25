# -*- coding: utf-8 -*-
from openpyxl import Workbook, load_workbook
import urllib
import json
# getting the right number of movies to correspond to actors
new_wb = Workbook()
new_ws = new_wb.active

old_wb = load_workbook(filename = 'actors.xlsx')
old_ws = old_wb['actors']

row_number = 1
is_cast_entry = 0
movie = ''
for row in old_ws.iter_rows('A{}:B{}'.format(old_ws.min_row,old_ws.max_row)):
    for cell in row:
    	if is_cast_entry == 0:
    		movie = cell.value
    		is_cast_entry = 1
    	else:
	    	arr = cell.value.split(', ')
	    	is_cast_entry = 0
	    	for entry in arr:
	    		if (entry != ''):
		    		new_ws.cell(row=row_number, column=1).value = movie
		    		row_number += 1
		    

new_wb.save('newactors.xlsx')